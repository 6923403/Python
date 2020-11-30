import requests
from bs4 import BeautifulSoup
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import json
import time
import threading
import pymysql
import smtplib
from email.mime.text import MIMEText
from email.header import Header 

"""
https://blog.csdn.net/XIAXU0101/article/details/104417413
https://www.jb51.net/article/200771.htm
https://www.cnblogs.com/mlp1234/p/9933919.html
https://www.jianshu.com/p/abb2d6e91c1f
"""

headers = {
    'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
}

class WangYi(object):
    def __init__(self, file_name):
        self.xlsxname = file_name

    def request(self):
        #new
        self.url = 'https://money.163.com/'
        self.data = requests.get(self.url, headers=headers)
        self.soup = BeautifulSoup(self.data.text, "lxml")

        self.time = time.time()

    def Style(self):
        self.m_font = Font(
        size = 12,
        bold = True,
        )

        self.head_font = Font(
        size = 14,
        bold = True,
        )

    """
    def Style(self):
        font = xlwt.Font()  # 内容字体
        font2 = xlwt.Font()  # 标题字体
        font3 = xlwt.Font()  # 指数
        font.height = 20 * 11
        font2.height = 20 * 12
        font2.bold = True
        font3.height = 20 * 13
        self.style = xlwt.XFStyle() #标题 链接字体
        self.style_head = xlwt.XFStyle() #类别列字体
        self.style_index = xlwt.XFStyle() #指数字体

        self.style.font = font
        self.style_head.font = font2
        self.style_index.font = font3
    """
    news = []
    def getTopNew(self):
        datalist = self.soup.select("ul li h2")
        """
        wb = Workbook()
        ws = wb['Sheet']
        wb.remove(ws)
        sheet = wb.create_sheet("Wy")
        
        t_row = 6
        t_col = 1
        sheet.cell(row=t_row, column=t_col, value="网易财经")
        t_row = t_row + 1
        sheet.cell(row=t_row, column=t_col, value="新闻标题")
        sheet.cell(row=t_row, column=t_col + 1, value="新闻链接")
        sheet.cell(row=t_row, column=t_col + 2, value="新闻简介")
        sheet.cell(row=t_row, column=t_col + 3, value="新闻时间")
        t_row = t_row + 1
        """

        conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='', db='test1', connect_timeout=1000)
        cursor = conn.cursor()
        for li in datalist:
            url = li.find('a')['href']
            title = li.get_text()
            self.news.append([title, url])
            sql_command = "insert into news(title, href) values(%s, %s)"
            cursor.execute(sql_command, (title, url))
            conn.commit()
        conn.close()

        from_addr = "***@163.com"
        passwd = "*******"
        mailhost = "smtp.163.com"

        to_addr = "***@qq.com"

        wy_mail = smtplib.SMTP()  # 建立SMTP对象
        wy_mail.connect(mailhost, 25)  # 25为SMTP常用端口
        wy_mail.login(from_addr, passwd)  # 登录邮箱

        content = ''
        for new in self.news:  # 拼接邮件内容字符串
            content += ' 新闻标题: ' + new[0] + '\n' + '新闻链接: ' + new[1] + '\n'
            content += '======================================================================\n'

        # 拼接题目字符串
        subject = time.strftime("%Y-%m-%d", time.localtime(time.time())) + "公司内部测试"

        # 加工邮件message格式
        msg = MIMEText(content, 'plain', 'utf-8')
        msg['From'] = "one<***@163.com>"
        msg['To'] = "two<***@qq.com>"
        msg['subject'] = Header(subject, 'utf-8')

        try:
            wy_mail.sendmail(from_addr, to_addr, msg.as_string())
            print('邮件发送成功')
        except Exception as e:
            print(str(e))
        wy_mail.quit()


        """
            sheet.cell(row=t_row, column=t_col, value=title)
            sheet.cell(row=t_row, column=t_col + 1, value=url)
            t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Wy Save Error = 1")

        """
    def getlist2(self):
        datalist2 = self.soup.find_all(class_='topnews_nlist topnews_nlist2')

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Wy')
        t_row = sheet.max_row # 获得行数
        t_col = 1
        for tp in datalist2:
            datalist3 = tp.select("li h3")
            for tn in datalist3:
                url = tn.find('a')['href']
                title = tn.get_text()
                sheet.cell(row=t_row, column=t_col, value=title)
                sheet.cell(row=t_row, column=t_col + 1, value=url)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("Wangyi Save Error = 2")

    def getstock(self):
        #stock
        stockurl = 'https://money.163.com/stock/'
        stockdata = requests.get(stockurl, headers=headers)
        soup = BeautifulSoup(stockdata.text, "lxml")
        stockl = soup.select('#stock2016_wrap > div > div.stock2016_content > div.idx_main.common_wrap.clearfix > div.news_main > div.news_main_wrap > div.topnews > div.topnews_first > h2 > a')
        top_url = stockl[0]['href']
        top_title = stockl[0].get_text()

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Wy')
        t_row = sheet.max_row  # 获得行数
        t_col = 1

        stocknewlist = soup.find_all(class_='topnews_list')
        for s_new in stocknewlist:
            news = s_new.find_all('a')
            for tn in news:
                t_url = tn['href']
                t_title = tn.get_text()
                sheet.cell(row=t_row, column=t_col, value=t_title)
                sheet.cell(row=t_row, column=t_col + 1, value=t_url)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("Wangyi Save Error = 4")

    threadLock = threading.Lock()
    def get_num(self, num, row, if_write):
        self.threadLock.acquire()

        num_price = num['price']
        num_open = num['open']
        num_updown = num['updown']
        num_high = num['high']
        num_low = num['low']
        num_yestclose = num['yestclose']
        num_percent = num['percent'] #num_percent 不直接写入
        num_update = num['update']
        m_percent = num_percent * 10000 #-0.020937 -%2.09
        percent = int(m_percent) /100

        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Wy')
        t_col = 1

        if if_write == True:
            t_row = 1
            sheet.cell(row=t_row, column=t_col, value='大盘指数')
            sheet.cell(row=t_row, column=t_col + 1, value='当前价位')
            sheet.cell(row=t_row, column=t_col + 2, value='今日涨幅')
            sheet.cell(row=t_row, column=t_col + 3, value='涨跌价格')
            sheet.cell(row=t_row, column=t_col + 4, value='开盘价位')
            sheet.cell(row=t_row, column=t_col + 5, value='今日最高')
            sheet.cell(row=t_row, column=t_col + 6, value='今日最低')
            sheet.cell(row=t_row, column=t_col + 7, value='昨日收盘')
            sheet.cell(row=t_row, column=t_col + 8, value='更新时间')
            sheet.cell(row=t_row + 2, column=t_col, value='深证成指')
            sheet.cell(row=t_row + 1, column=t_col, value='上证指数')
            sheet.cell(row=t_row + 3, column=t_col, value='沪深300')
            t_row = t_row + 1

        sheet.cell(row=row, column=2, value=num_price)
        sheet.cell(row=row, column=3, value=str(percent) + "%")
        sheet.cell(row=row, column=4, value=num_updown)
        sheet.cell(row=row, column=5, value=num_open)
        sheet.cell(row=row, column=6, value=num_high)
        sheet.cell(row=row, column=7, value=num_low)
        sheet.cell(row=row, column=8, value=num_yestclose)
        sheet.cell(row=row, column=9, value=num_update)
        try:
            wb.save(self.xlsxname)
            self.threadLock.release()
        except:
            self.threadLock.release()
            print("Wangyi Save Error = 5")


    def getindex(self):
        #index
        indexurl = 'http://api.money.126.net/data/feed/1399001,1399300,0000001,HSRANK_COUNT_SHA,HSRANK_COUNT_SZA,HSRANK_COUNT_SH3?callback=ne_{}&[object%20Object]'.format(int(self.time))
        indexdata = requests.get(indexurl, headers=headers)
        #soup = BeautifulSoup(self.indexdata.text,)
        data = indexdata.text + "del"
        time = int(self.time)
        data = data.replace('ne_' + str(time) + '(', '')
        data = data.replace(');del', '')
        #json_d = json.dumps(data)  # 编码
        json_str = json.loads(data)  # 解码

        if_write = True
        n_data1 = json_str['0000001'] #上证指数_0000001
        n_data2 = json_str['1399001'] #深证成指_1399001
        n_data3 = json_str['1399300'] #沪深300_1399300
        t1 = threading.Thread(target=self.get_num, args=(n_data1, 2, if_write, ))
        t1.start()
        if_write = False
        t1.join()
        t2 = threading.Thread(target=self.get_num, args=(n_data2, 3, if_write, ))
        t3 = threading.Thread(target=self.get_num, args=(n_data3, 4, if_write, ))
        t2.start()
        t3.start()
        t2.join()
        t3.join()
        #self.get_num(n2, 2, if_write)
        #self.get_num(n3, 3, if_write)

    def get_bu(self, soup, if_write):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Wy')
        t_row = sheet.max_row + 1
        t_col = 1

        if if_write == True:
            sheet.cell(row=t_row + 1, column=t_col, value="市场资讯")
            sheet.cell(row=t_row + 2, column=t_col, value="新闻标题")
            sheet.cell(row=t_row + 2, column=t_col + 1, value="新闻链接")
            sheet.cell(row=t_row + 2, column=t_col + 2, value="新闻简介")
            sheet.cell(row=t_row + 2, column=t_col + 3, value="新闻时间")
            t_row = t_row + 3  # 已经使用多少行

        datalist1 = soup.find_all(class_='list_item clearfix')
        for Newslist in datalist1:
            News = Newslist.find_all(class_='item_top')
            for m_new in News:
                m_new1 = m_new.find('a')
                m_new2 = m_new.find(class_='time')
                m_title = m_new1.get_text()
                m_url = m_new1['href']
                m_time = m_new2.get_text()
                sheet.cell(row=t_row, column=t_col, value=m_title)
                sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                sheet.cell(row=t_row, column=t_col + 3, value=m_time)
                t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except:
            print("Wangyi Save Error = 6")


    def getBusiness(self): #市场资讯 获取两页
        bu_url = 'http://money.163.com/special/00251LR5/cpznList.html'
        bu_url2 = 'http://money.163.com/special/00251LR5/cpznList_02.html' #第二页
        
        bu_data1 = requests.get(bu_url, headers=headers)
        soup1 = BeautifulSoup(bu_data1.text, "lxml")

        bu_data2 = requests.get(bu_url2, headers=headers)
        soup2 = BeautifulSoup(bu_data2.text, "lxml")

        if_write = True
        t1 =threading.Thread(target=self.get_bu, args=(soup1, if_write, ))
        t1.start()
        if_write = False
        t1.join()
        t2 =threading.Thread(target=self.get_bu, args=(soup1, if_write, ))
        t2.start()
        #self.get_bu(soup2, if_write)


    def get_Indu(self, soup, if_write):
        wb = load_workbook(self.xlsxname)
        sheet = wb.get_sheet_by_name('Wy')
        t_row = sheet.max_row + 1
        t_col = 1
        if if_write == True:
            sheet.cell(row=t_row + 1, column=t_col, value="行业板块")
            sheet.cell(row=t_row + 2, column=t_col, value="新闻标题")
            sheet.cell(row=t_row + 2, column=t_col + 1, value="新闻链接")
            sheet.cell(row=t_row + 2, column=t_col + 2, value="新闻简介")
            sheet.cell(row=t_row + 2, column=t_col + 3, value="新闻时间")
            t_row = t_row + 3

        datalist = soup.find_all(class_="col_l")
        for Newslist in datalist:
            News = Newslist.find_all(class_="list_item clearfix")
            for newlist in News:
                news = newlist.find_all(class_="item_top")
                for new in news:
                    m_new = new.select('h2 a')
                    m_url = m_new[0]['href']
                    m_title = m_new[0].get_text()
                    m_new2 = new.select('p span')
                    m_time = m_new2[0].get_text()
                    sheet.cell(row=t_row, column=t_col, value=m_title)
                    sheet.cell(row=t_row, column=t_col + 1, value=m_url)
                    sheet.cell(row=t_row, column=t_col + 3, value=m_time)
                    t_row = t_row + 1
        try:
            wb.save(self.xlsxname)
        except Exception:
            print("Wangyi Save Error = 7")

    def getIndustry(self): #行业资讯 前两页
        url = 'http://money.163.com/special/00251LJV/hyyj.html'
        url2 = 'http://money.163.com/special/00251LJV/hyyj_02.html'

        Industry_data1 = requests.get(url, headers=headers)
        soup1 = BeautifulSoup(Industry_data1.text, "lxml")
        Industry_data2 = requests.get(url2, headers=headers)
        soup2 = BeautifulSoup(Industry_data2.text, "lxml")
        if_write = True
        t1 = threading.Thread(target=self.get_Indu, args=(soup1, if_write, ))
        t1.start()
        if_write = False
        t1.join()
        t2 = threading.Thread(target=self.get_Indu, args=(soup2, if_write, ))
        t2.start()
        t2.join()
        #self.get_Indu(soup2, if_write)


    def main(self, file_name):
        Wy = WangYi(file_name)
        Wy.Style()
        Wy.request()
        Wy.getTopNew()
        """
        Wy.getlist2()
        #stock
        Wy.getstock()
        #Wy.getindex()  # 主页原创栏目右边
        Wy.getBusiness()
        Wy.getIndustry()
        """
