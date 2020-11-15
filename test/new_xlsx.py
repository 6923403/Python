from openpyxl import Workbook
from openpyxl.styles import Font
filename = './Finance/Finance.xlsx'

wb = Workbook()
ws = wb['Sheet']
wb.remove(ws)
sheet = wb.create_sheet("Wy")
sheet.cell(row = 1, column = 1, value = "测试")

wb.save(filename)
